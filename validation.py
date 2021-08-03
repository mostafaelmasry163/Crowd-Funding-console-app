import os
import openpyxl
import datetime

projectsdatabook = openpyxl.load_workbook("projects_data.xlsx")
projectsdata = projectsdatabook["Sheet1"]

class Validation :
    def titlevalidation(title):
        j = 1
        while (not title) or (projectsdata.cell(j, 1).value is not None):
            if not title:
                print("title should have at least one character")
                j = 1
                title = input("please, enter project title :  ")
            elif projectsdata.cell(j, 1).value == title:
                print("this title is already exist")
                j = 1
                title = input("please, enter project title :  ")
            else:
                j += 1
        os.system('clear')
        return title

    def targetvalidation(target):
        while (not target) or (target.isdigit() is False):
            if not target:
                print("target should have at least one digit")
                target = input("please, enter project total target :  ")
            elif target.isdigit() is False:
                print("project target can be only digits")
                target = input("please, enter project total target :  ")
        os.system('clear')
        return target

    def startdatevalidation(date):
        condition = True
        while condition:
            try:
                start_date_ob = datetime.datetime.strptime(
                    date, "%d/%m/%Y")
                condition = False
                if start_date_ob.date() < datetime.datetime.now().date():
                    print("start date cannot be a previous date to today date ")
                    date = input(
                        "please, enter project start date in form [Day/Month/Year] :  ")
                    condition = True
            except:
                print("invalid formula")
                date = input(
                    "please, enter project start date in form [Day/Month/Year] :  ")
        os.system('clear')
        return date

    def endDatevalidation(start,end):
        condition2 = True
        while condition2:
            try:
                start_date_ob = datetime.datetime.strptime(
                    start, "%d/%m/%Y")
                end_date_ob = datetime.datetime.strptime(
                    end, "%d/%m/%Y")
                condition2 = False
                if end_date_ob.date() < start_date_ob.date():
                    print("end date cannot be a previous date to project start date ")
                    end = input(
                        "please, enter project end date in form [Day/Month/Year] :  ")
                    condition2 = True
            except:
                print("invalid formula")
                end = input(
                    "please, enter project end date in form [Day/Month/Year] :  ")
        os.system('clear')
        return end
