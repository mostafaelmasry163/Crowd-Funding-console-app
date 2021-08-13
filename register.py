import re
from openpyxl import load_workbook

filename = "users.xlsx"
wb = load_workbook(filename)
ws = wb.worksheets[0]


class User:
    def __init__(self):
        self.email = ""

    def Register_new_user(self):
        print(">>>>Hello New Crowdfunder<<<<")
        self.get_user_fname()

    def get_user_fname(self):
        global fname
        fname = input("Enter your First name:")
        if not fname or fname.isdecimal():
            print(">>>Please enter valid user name<<<")
            self.get_user_fname()
        else:
            self.get_user_lname()
        return

    def get_user_lname(self):
        global lname
        lname = input("Enter your Last name:")
        if not lname or lname.isdecimal():
            print(">>>Please enter valid user name<<<")
            self.get_user_lname()
        else:
            self.get_user_email()
        return

    def get_user_email(self):
        regex_email = r'\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
        global email
        self.email = input("Enter your Email:")
        j = 1
        Flag = 0
        while (ws.cell(j, 3).value is not None):
            if ws.cell(j, 3).value == self.email:
                print("This email is already registered")
                Flag += 1
                self.get_user_email()
                break
            elif not (re.match(regex_email, self.email)):
                print(">>>Enter a valid Email<<<")
                Flag += 1
                self.get_user_email()
                break
            else:
                j += 1
        if (Flag < 1):
            self.get_user_password()
        return

    def get_user_password(self):
        global password
        try:
            password = int(input("Enter your Password:"))
            confirm_password = int(input("Confirm Password:"))
            if confirm_password != password:
                print(">>>Password does not match<<<")
                self.get_user_password()
            else:
                self.get_user_mobile()
        except:
            print(">>>Please enter password<<<")
            self.get_user_password()
        return

    def get_user_mobile(self):
        global mobile
        mobile = input("Enter your Mobile phone:")
        if (self.isValid(mobile)):
            print("---User Successfuly Added---")
            i = 1
            while ws.cell(i, 5).value is not None:
                i += 1
            ws.cell(i, 1).value = fname
            ws.cell(i, 2).value = lname
            ws.cell(i, 3).value = self.email
            ws.cell(i, 4).value = password
            ws.cell(i, 5).value = mobile
            wb.save(filename)
        else:
            print(">>>please add valid egyption phone number<<<")
            self.get_user_mobile()
        return

    def isValid(self, s):
        Pattern = re.compile("(01)?[0-2][0-9]{8}")
        return Pattern.match(s)

    # login starts from here
    def login(self):
        self.email = input("please enter your email to login : ")
        password = str(input("please enter your password to login : "))

        for row in ws.iter_rows(min_row=2, max_row=100, min_col=3, max_col=4, values_only=True):
            if row[0] == self.email and str(row[1]) == password:
                print("login successfully")
                break
                # take action function call
        else:
            print("Invalid email or password")
            x = int(input("To try again press (1)   For Registration press (2)   "))
            if x == 1:
                self.login()
            if x == 2:
                self.get_user_fname()
