import validation
import openpyxl
import os


projectsdatabook = openpyxl.load_workbook("projects_data.xlsx")
projectsdata = projectsdatabook["Sheet1"]


class Options :
    # View All Projects
    def findAllProject():
        allProjectsDictionry = {}
        for project_row in range(2, projectsdata.max_row+1):
            for project_col in range(1,6):
                projectDictionry = {
                    'title': projectsdata.cell(project_row,1).value,
                    'details': projectsdata.cell(project_row,2).value,
                    'total': projectsdata.cell(project_row,3).value,
                    'startDate': projectsdata.cell(project_row,4).value,
                    'endDate': projectsdata.cell(project_row,5).value
                }
            allProjectsDictionry[projectDictionry['title']] = projectDictionry
        return allProjectsDictionry

    # create project
    def create(self):
        # clearing console
        os.system('clear')
        # Open File
        projectsdatabook = openpyxl.load_workbook("projects_data.xlsx")
        projectsdata = projectsdatabook["Sheet1"]

        # Getting project title and make sure that there is no similar title and it is not empty
        title = input("please, enter project title :  ")
        title = validation.Validation.titlevalidation(title)

        # Getting project details
        details = input("please, enter project details :  ")
        os.system('clear')

        # Get and validate project total target
        total_target = input("please, enter project total target :  ")
        total_target = validation.Validation.targetvalidation(total_target)

        # Get campaign start date
        user_start_date = input(
            "please, enter project start date in form [Day/Month/Year] :  ")
        user_start_date = validation.Validation.startdatevalidation(user_start_date)       

        # Get campaign start date
        user_end_date = input(
            "please, enter project end date in form [Day/Month/Year] :  ")
        user_end_date = validation.Validation.endDatevalidation(user_start_date, user_end_date)

        # showing project data
        os.system('clear')
        print(f'your project information is : \n'
              f'Project title : {title}  \n'
              f'Project details : {details}  \n'
              f'Project total target : {total_target}  \n'
              f'Project start date and end date :  {user_start_date}  ,  {user_end_date}')

        # writing data
        i = projectsdata.max_row+1
        while projectsdata.cell(i, 6).value is not None:
            i += 1
        # projectsdata.cell(i, 6).value = self.user_mail
        projectsdata.cell(i, 1).value = title
        projectsdata.cell(i, 2).value = details
        projectsdata.cell(i, 3).value = total_target
        projectsdata.cell(i, 4).value = user_start_date
        projectsdata.cell(i, 5).value = user_end_date
        projectsdatabook.save("projects_data.xlsx")

        print("End of creation")
