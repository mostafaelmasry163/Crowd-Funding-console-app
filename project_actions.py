import openpyxl
import os
import datetime


class Project:
    
    def new_action(self, email):
        try:
            decision = int(input(f'please choose what do you want to do : \n '
                                 f'Enter 1 for creating new project.\n'
                                 f' Enter 2 for viewing all projects. \n'
                                 f' Enter 3 for viewing all your projects. \n'
                                 f' Enter 4 for edit your projects. \n'
                                 f' Enter 5 for deleting one of your projects.\n'
                                 f' Enter 6 for searching a project by its date.\n'
                                 f' type here : '))
        except:
            print("invalid, try again ")
            self.new_action(email)
        else:
            if decision == 1:
                #creat new project
                self.create(email)
            elif decision == 2:
                # viewing all projects
                print(f'All projects informations \n')
                self.view()
            elif decision == 3:
                # filter view by user email(projects by the user)
                self.view(email)
            elif decision == 4:
                #edit project
                project_name = input(
                    "please enter the project name you want to edit: ")
                self.edit(project_name, email)
            elif decision == 5:
                # delete project
                project_name = input(
                    "please enter the project name you want to delete: ")
                self.delete(project_name,email)
            elif decision == 6:
                # Get search date
                user_search_date = input(
                    "please, enter project start date you want to search in form [Day/Month/Year] :  ")
                condition = True
                while condition:
                    try:
                        start_date_ob = datetime.datetime.strptime(
                            user_search_date, "%d/%m/%Y")
                        condition = False
                    except:
                        print("invalid formula")
                        user_search_date = input(
                            "please, enter project start date in form [Day/Month/Year] :  ")
                # filter view to show specific date
                self.view(email, user_search_date)
            else:
                print("invalid, try again ")
                self.new_action(email)

    def create(self, email, indx=0):
        # Open File
        my_workbook = openpyxl.load_workbook("projects_data.xlsx")
        projects_sheet = my_workbook["Sheet1"]
        # clearing console
        os.system('clear')
        # Getting project title and make sure that there is no similar title and it is not empty
        j = 1
        title = input("please, enter project title :  ")
        while (not title) or (projects_sheet.cell(j, 1).value is not None):
            if not title:
                print("title should have at least one character")
                j = 1
                title = input("please, enter project title :  ")
            elif projects_sheet.cell(j, 1).value == title:
                print("this title is already exist")
                j = 1
                title = input("please, enter project title :  ")
            else:
                j += 1
        os.system('clear')

        # Getting project details
        details = input("please, enter project details :  ")
        os.system('clear')

        # Get and validate project total target
        total_target = input("please, enter project total target :  ")
        while (not total_target) or (total_target.isdigit() is False):
            if not total_target:
                print("target should have at least one digit")
                total_target = input("please, enter project total target :  ")
            elif total_target.isdigit() is False:
                print("project target can be only digits")
                total_target = input("please, enter project total target :  ")
        os.system('clear')

        # Get campaign start date
        user_start_date = input(
            "please, enter project start date in form [Day/Month/Year] :  ")
        condition = True
        while condition:
            try:
                start_date_ob = datetime.datetime.strptime(
                    user_start_date, "%d/%m/%Y")
                condition = False
                if start_date_ob.date() < datetime.datetime.now().date():
                    print("start date cannot be a previous date to today date ")
                    user_start_date = input(
                        "please, enter project start date in form [Day/Month/Year] :  ")
                    condition = True
            except:
                print("invalid formula")
                user_start_date = input(
                    "please, enter project start date in form [Day/Month/Year] :  ")
        os.system('clear')

        # Get campaign end date
        user_end_date = input(
            "please, enter project end date in form [Day/Month/Year] :  ")
        condition2 = True
        while condition2:
            try:
                end_date_ob = datetime.datetime.strptime(
                    user_end_date, "%d/%m/%Y")
                condition2 = False
                if end_date_ob.date() < start_date_ob.date():
                    print("end date cannot be a previous date to project start date ")
                    user_end_date = input(
                        "please, enter project end date in form [Day/Month/Year] :  ")
                    condition2 = True
            except:
                print("invalid formula")
                user_end_date = input(
                    "please, enter project end date in form [Day/Month/Year] :  ")
        os.system('clear')

        # showing project data
        os.system('clear')
        print(f'your project information is : \n'
              f'Project title : {title}  \n'
              f'Project details : {details}  \n'
              f'Project total target : {total_target}  \n'
              f'Project start date and end date :  {user_start_date}  ,  {user_end_date}')

        # writing data
        # i = projects_sheet.max_row
        if indx == 0 :
            indx = 1
            while projects_sheet.cell(indx, 6).value is not None:
                indx += 1
            self.write_in_xl(indx,title,details,total_target,user_start_date,user_end_date,email)
            print("End of creation")
        else :
            self.write_in_xl(indx, title, details, total_target,user_start_date, user_end_date, email)
  
    def view(self, email='', date=''):
        # Open File
        my_workbook = openpyxl.load_workbook("projects_data.xlsx")
        projects_sheet = my_workbook["Sheet1"]
        
        if (email == '') & (date == ''):
            for project_row in range(2, projects_sheet.max_row+1):
                title = projects_sheet.cell(project_row, 1).value
                details = projects_sheet.cell(project_row, 2).value
                total = projects_sheet.cell(project_row, 3).value
                startdate = projects_sheet.cell(project_row, 4).value
                enddate = projects_sheet.cell(project_row, 5).value
                owner = projects_sheet.cell(project_row, 6).value
                print(f'{title} project information is : \n'
                      f'Project owner : {owner}  \n'
                      f'Project title : {title}  \n'
                      f'Project details : {details}  \n'
                      f'Project total target : {total}  \n'
                      f'Project start date and end date:  {startdate}  ,  {enddate}\n'
                      f'--------------------------------  \n')
        elif (date != '') & (email != ''):
            print(f'All projects that starts in {date} \n')
            allprojectsdates = []
            for project_row in range(2, projects_sheet.max_row+1):
                allprojectsdates.append(
                    projects_sheet.cell(project_row, 4).value)
            if date not in allprojectsdates:
                return print("date not found")
            i = 0
            while i < len(allprojectsdates):
                if date == allprojectsdates[i]:
                    title = projects_sheet.cell(i+2, 1).value
                    details = projects_sheet.cell(i+2, 2).value
                    total = projects_sheet.cell(i+2, 3).value
                    startdate = projects_sheet.cell(i+2, 4).value
                    enddate = projects_sheet.cell(i+2, 5).value
                    owner = projects_sheet.cell(i+2, 6).value
                    print(f'{title} project information is : \n'
                          f'Project owner : {owner}  \n'
                          f'Project title : {title}  \n'
                          f'Project details : {details}  \n'
                          f'Project total target : {total}  \n'
                          f'Project start date and end date:  {startdate}  ,  {enddate}\n'
                          f'--------------------------------  \n')
                    i += 1
                else:
                    i += 1
        elif (email != '') & (date ==''):
            print(f'All your projects \n')
            allprojectsemails = []
            for project_row in range(2, projects_sheet.max_row+1):
                allprojectsemails.append(
                    projects_sheet.cell(project_row, 6).value)
            if email not in allprojectsemails:
                return print("you don't have any projects")
            i = 0
            while i < len(allprojectsemails):
                if email == allprojectsemails[i]:
                    title = projects_sheet.cell(i+2, 1).value
                    details = projects_sheet.cell(i+2, 2).value
                    total = projects_sheet.cell(i+2, 3).value
                    startdate = projects_sheet.cell(i+2, 4).value
                    enddate = projects_sheet.cell(i+2, 5).value
                    owner = projects_sheet.cell(i+2, 6).value
                    print(f'{title} project information is : \n'
                          f'Project owner : You  \n'
                          f'Project title : {title}  \n'
                          f'Project details : {details}  \n'
                          f'Project total target : {total}  \n'
                          f'Project start date and end date:  {startdate}  ,  {enddate}\n'
                          f'--------------------------------  \n')
                    i += 1
                else:
                    i += 1

    def delete(self, name, email): 
        # Open File
        my_workbook = openpyxl.load_workbook("projects_data.xlsx")
        projects_sheet = my_workbook["Sheet1"]       
        allprojectsnames = []
        allprojectsemails = []
        for project_row in range(2, projects_sheet.max_row+1):
            allprojectsnames.append(projects_sheet.cell(project_row, 1).value)
            allprojectsemails.append(
                projects_sheet.cell(project_row, 6).value)
        if (email in allprojectsemails) & (name in allprojectsnames):
            if email == allprojectsemails[allprojectsnames.index(name)]:
                answer = input(
                f'Are you sure you want to delete {name} project?(Y/N): ')
                if answer.upper() == 'Y':
                    projects_sheet.delete_rows(allprojectsnames.index(name)+2)
                    my_workbook.save("projects_data.xlsx")
                    print("project deleted")
                elif answer.upper() == 'N':
                    return 
            else :
                print("you don't own this project")
        else :
            print("project name is not found")

    def edit(self, name,email):
        # Open File
        my_workbook = openpyxl.load_workbook("projects_data.xlsx")
        projects_sheet = my_workbook["Sheet1"]
        allprojectsnames = []
        allprojectsemails = []
        for project_row in range(2, projects_sheet.max_row+1):
            allprojectsnames.append(projects_sheet.cell(project_row, 1).value)
            allprojectsemails.append(
                projects_sheet.cell(project_row, 6).value)
        if (email in allprojectsemails) & (name in allprojectsnames):
            if email == allprojectsemails[allprojectsnames.index(name)]:
                answer = input(
                f'Are you sure you want to edit {name} project?(Y/N): ')
                if answer.upper() == 'Y':
                    self.create(email, allprojectsnames.index(name)+2)
                    print("project edited")
                elif answer.upper() == 'N':
                    return
            else:
                print("you don't own this project")
        else:
            print("project name is not found")

    def write_in_xl(self, indx,title,details,total_target,user_start_date,user_end_date,email):
        # Open File
        my_workbook = openpyxl.load_workbook("projects_data.xlsx")
        projects_sheet = my_workbook["Sheet1"]
        indx=int(indx)
        projects_sheet.cell(indx, 1).value = title
        projects_sheet.cell(indx, 2).value = details
        projects_sheet.cell(indx, 3).value = total_target
        projects_sheet.cell(indx, 4).value = user_start_date
        projects_sheet.cell(indx, 5).value = user_end_date
        projects_sheet.cell(indx, 6).value = email
        my_workbook.save("projects_data.xlsx")
        